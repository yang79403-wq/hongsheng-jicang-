#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""洪盛集藏本地知识库构建器：PDF/Word/Excel/图片/文本 -> 标准资料包。
不自动发布，先输出 data/local-knowledge/staging.json。"""
import pathlib,json,hashlib,datetime,os
ROOT=pathlib.Path(os.environ.get('HONGSHENG_LIBRARY','洪盛资料库')); OUT=pathlib.Path('data/local-knowledge'); OUT.mkdir(parents=True,exist_ok=True)

def text_file(p):
 b=p.read_bytes()
 for e in ('utf-8','gb18030','utf-16'):
  try:return b.decode(e)
  except:pass
 return ''
def extract(p):
 ext=p.suffix.lower()
 if ext in {'.txt','.md','.csv'}: return text_file(p),[]
 if ext=='.json': return text_file(p),[]
 if ext=='.pdf':
  try:
   from pypdf import PdfReader
   r=PdfReader(str(p)); return '\n'.join(x.extract_text() or '' for x in r.pages),[]
  except Exception as e:return '',['PDF解析失败: '+str(e)]
 if ext=='.docx':
  try:
   from docx import Document
   d=Document(str(p)); return '\n'.join(x.text for x in d.paragraphs),[]
  except Exception as e:return '',['Word解析失败: '+str(e)]
 if ext in {'.xlsx','.xlsm'}:
  try:
   from openpyxl import load_workbook
   w=load_workbook(str(p),read_only=True,data_only=True); rows=[]
   for s in w.worksheets:
    rows.append('【工作表：'+s.title+'】')
    for r in s.iter_rows(values_only=True): rows.append(' | '.join('' if v is None else str(v) for v in r))
   return '\n'.join(rows),[]
  except Exception as e:return '',['Excel解析失败: '+str(e)]
 return '',[]
def sha(p):return hashlib.sha256(p.read_bytes()).hexdigest()
def classify(p):
 names={'福建钱币':('fujian','福建铜钱 银元 铜币 纸币'),'评级资料':('services','评级知识'),'行情资料':('market','行情资料参考')}
 for k,v in names.items():
  if k in p.parts:return v
 return ('content','收藏交流 · 鉴赏参考')
items=[]
for p in ROOT.rglob('*') if ROOT.exists() else []:
 if not p.is_file() or p.name.startswith('_') or p.suffix.lower() not in {'.txt','.md','.json','.csv','.pdf','.docx','.xlsx','.xlsm','.jpg','.jpeg','.png','.webp'}:continue
 txt,errors=extract(p); gid,gname=classify(p); items.append({'id':'local_'+sha(p)[:16],'title':p.stem,'sourceFile':str(p),'sourceHash':sha(p),'date':datetime.date.today().isoformat(),'groupId':gid,'groupName':gname,'content':txt,'images':([str(p)] if p.suffix.lower() in {'.jpg','.jpeg','.png','.webp'} else []),'status':'draft','needsReview':True,'parseErrors':errors})
(OUT/'staging.json').write_text(json.dumps(items,ensure_ascii=False,indent=2),encoding='utf-8')
print('已建立本地知识库暂存：',OUT/'staging.json',' 共 ',len(items),' 条；默认草稿，未发布。')
